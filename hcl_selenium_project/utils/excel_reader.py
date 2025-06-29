from openpyxl import load_workbook

def read_credentials(file_path):
    wb = load_workbook(file_path)
    ws = wb.active
    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        data.append(row)
    return data