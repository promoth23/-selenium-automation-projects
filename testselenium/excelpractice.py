from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from openpyxl import load_workbook
from webdriver_manager.chrome import ChromeDriverManager
import time

# Load Excel file
wb = load_workbook("Book1.xlsx")
ws = wb.active

# Launch browser
driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))
driver.maximize_window()

# Loop through each row in Excel (skip header)
for row in ws.iter_rows(min_row=2, values_only=True):
    username_value, password_value = row
    print(f"\n🔁 Testing login with: {username_value} / {password_value}")

    # Open login page
    driver.get("https://practicetestautomation.com/practice-test-login/")
    time.sleep(2)

    # Fill the login form
    driver.find_element(By.ID, "username").send_keys(username_value)
    driver.find_element(By.ID, "password").send_keys(password_value)
    driver.find_element(By.ID, "submit").click()
    time.sleep(2)

    # Check if login is successful
    if "Logged In Successfully" in driver.page_source:
        print("✅ Login successful!")
    else:
        print("❌ Login failed!")

# Quit browser after all tests
driver.quit()





